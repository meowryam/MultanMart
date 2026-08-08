"""Rebuild products.json from the PIVOT TABLE worksheet only.

The PIVOT TABLE sheet in the inventory workbook is the single source of truth
for the product catalogue:

  * price  <- "TP" column
  * stock  <- "QTY" column

Rules:
  * Every SKU present in the pivot is written to products.json.
  * Products already in products.json whose SKU is NOT in the pivot are removed.
  * Existing metadata (name/category/brand/image) is preserved for products
    that are already in products.json so the frontend keeps working.
  * New pivot rows (SKUs never seen before) are added with cleaned names,
    a best-guess category/brand, and an image slug.
  * The "Grand Total" row is skipped; duplicate SKU rows are merged.

Usage (from the project root):

    python tools/rebuild_products.py [path/to/inventory.xlsx]

Output: products.json in the project root.
"""

import json
import os
import re
import sys

import openpyxl

DEFAULT_INPUT = "stock_in_hand 1.xlsx"
SHEET_NAME = "PIVOT TABLE"
HEADER_ROW = 3  # header is row 3, data starts at row 4

# Column indexes (0-based) in the PIVOT TABLE worksheet.
COL = {
    "brand": 1,   # "Brand "
    "sku": 2,     # "SKU"
    "desc": 3,    # "SKU Description"
    "qty": 4,     # "QTY"
    "tp": 5,      # "TP"
}

DC_COUNTER_SKU_RE = re.compile(r"^DC\d+$", re.IGNORECASE)
DC_COUNTER_LABEL = "DC COUNTER"

# Anything that looks like a price embedded in a description/name, e.g.
# "Rs 170", "RS:50", "RS100", "Rs.240", "PKR 500". These are OLD prices and
# must never leak into product names — the real price lives in the TP column.
PRICE_TOKEN_RE = re.compile(
    r"\b(?:Rs\.?|PKR\.?)\s*:?\s*\d+\b", re.IGNORECASE
)
CP_TOKEN_RE = re.compile(r"\bCP\s*\d*\b", re.IGNORECASE)
SPACE_RE = re.compile(r"\s{2,}")

# Best-guess (category, brand) for pivot SKUs that were never in products.json.
# These are the only rows without existing website metadata.
NEW_PRODUCT_META = {
    "8964002548545": ("skin-care", "SKIN CARE PRODUCTS"),   # KESHIA ALMOND OIL
    "580": ("stationery", ""),                              # TRANSPARENT TAPE
    "8886950099477": ("home-care", "LEMON MAX"),            # MAX LIQ 170ML
    "8886950039381": ("skin-cleansing", "PALMOLIVE"),       # PALMOLIVE TRY PACK
    "8964000020197": ("spices", "SHANGRILA SAUCE"),         # SHANGRILA CHILI SAUCE 700ML
    "8964000020098": ("spices", "SHANGRILA SAUCE"),         # SHANGRILA SOYA SAUCE 700ML
    "8961014034806": ("spices", "SHANGRILA SAUCE"),         # CHILI G 190G
    "8961014266153": ("hair-care", "LIFE BUOY SHAMPOO"),    # LBS ONION 650ML
    "8961014263732": ("skin-cleansing", "LIFE BUOY SOAP"),  # LB CARE 70G
}


def slugify(value):
    """Turn a product name into a kebab-case image filename (no extension)."""
    text = re.sub(r"[^A-Za-z0-9]+", "-", value)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text.lower() or "product"


def clean_name(raw):
    """Strip price/consumer-price text from a pivot description."""
    name = str(raw or "").strip()
    name = PRICE_TOKEN_RE.sub(" ", name)
    name = CP_TOKEN_RE.sub(" ", name)
    name = SPACE_RE.sub(" ", name).strip(" -")
    return name


def strip_prices(name):
    """Sanitize a final product name: drop any embedded price tokens."""
    name = PRICE_TOKEN_RE.sub(" ", name)
    name = CP_TOKEN_RE.sub(" ", name)
    name = SPACE_RE.sub(" ", name).strip(" -")
    return name


def clean_brand(raw):
    """Return '' for the pivot's DC COUNTER group label, else the raw value."""
    brand = str(raw or "").strip()
    return "" if brand.upper() == DC_COUNTER_LABEL else brand


def norm_price(raw):
    """Return TP as an int when whole, else a float rounded to 2 decimals."""
    value = float(raw)
    return int(value) if value.is_integer() else round(value, 2)


def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_full = os.path.join(project_root, input_path)

    workbook = openpyxl.load_workbook(input_full, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {input_full}.")
    ws = workbook[SHEET_NAME]

    # Read the pivot in sheet order: SKU -> record (first occurrence wins).
    pivot = []
    seen_skus = set()
    duplicate_rows = 0
    skipped_rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        sku_raw = row[COL["sku"]]
        if sku_raw is None or str(sku_raw).strip() == "":
            skipped_rows.append(("blank SKU", row))
            continue
        sku = str(sku_raw).strip()
        desc = str(row[COL["desc"]] or "").strip()
        if desc == "Grand Total":
            skipped_rows.append(("grand total", row))
            continue
        if DC_COUNTER_SKU_RE.match(sku):
            skipped_rows.append(("dc counter", row))
            continue
        if sku in seen_skus:
            duplicate_rows += 1
            skipped_rows.append(("duplicate SKU", row))
            continue
        seen_skus.add(sku)
        pivot.append(
            {
                "sku": sku,
                "desc": desc,
                "brand": clean_brand(row[COL["brand"]]),
                "qty": row[COL["qty"]],
                "tp": row[COL["tp"]],
            }
        )

    json_path = os.path.join(project_root, "products.json")
    with open(json_path, "r", encoding="utf-8") as fh:
        existing = json.load(fh)
    by_sku = {str(p["sku"]).strip(): p for p in existing}

    products = []
    added = 0
    removed = 0
    price_changed = 0
    stock_changed = 0
    name_cleaned = 0

    for pv in pivot:
        old = by_sku.get(pv["sku"])
        if old is not None:
            product = dict(old)
            if "brand" in product:
                product["brand"] = clean_brand(product["brand"])
            if norm_price(pv["tp"]) != old["price"]:
                price_changed += 1
            if int(pv["qty"]) != old["stock"]:
                stock_changed += 1
        else:
            category, brand = NEW_PRODUCT_META.get(pv["sku"], ("others", ""))
            name = clean_name(pv["desc"]) or pv["brand"] or pv["sku"]
            product = {
                "name": name,
                "category": category,
                "brand": brand or pv["brand"],
                "sku": pv["sku"],
                "image": f"images/{slugify(name)}.webp",
            }
            added += 1
        clean = strip_prices(product["name"])
        if clean != product["name"]:
            name_cleaned += 1
        product["name"] = clean
        product["price"] = norm_price(pv["tp"])
        product["stock"] = int(pv["qty"])
        products.append(product)

    existing_skus = set(by_sku)
    pivot_skus = set(seen_skus)
    removed = len(existing_skus - pivot_skus)

    next_id = (max((p["id"] for p in products if "id" in p), default=0) + 1)
    for p in products:
        if "id" not in p:
            p["id"] = next_id
            next_id += 1

    # Match the existing file's CRLF formatting (no trailing newline).
    body = json.dumps(products, ensure_ascii=False, indent=2).replace("\n", "\r\n")
    with open(json_path, "w", encoding="utf-8", newline="") as fh:
        fh.write(body)

    print(f"Pivot product rows:      {len(pivot)}")
    print(f"Duplicate rows merged:   {duplicate_rows}")
    print(f"Skipped rows:            {len(skipped_rows) - duplicate_rows}")
    for kind, row in skipped_rows:
        print(f"   - {kind}: SKU={row[COL['sku']]} desc={row[COL['desc']]!r} qty={row[COL['qty']]} tp={row[COL['tp']]}")
    print(f"Products written:        {len(products)}")
    print(f"Products added:          {added}")
    print(f"Products removed:        {removed}")
    print(f"Prices updated:          {price_changed}")
    print(f"Quantities updated:      {stock_changed}")
    print(f"Names cleaned:           {name_cleaned}")


if __name__ == "__main__":
    main()
