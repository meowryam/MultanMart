"""One-time inventory import.

Reads the Excel inventory file (sheet: stock_in_hand) and writes products.json,
which the website loads at runtime. The website never reads the Excel file
directly — re-run this script to refresh products.json after editing stock.

Usage (from the project root):

    python tools/import_inventory.py [path/to/inventory.xlsx]

Output: products.json in the project root.
"""

import json
import os
import re
import sys

import openpyxl

DEFAULT_INPUT = "stock_in_hand.xlsx"
SHEET_NAME = "stock_in_hand"
HEADER_ROW = 11  # first data row is HEADER_ROW + 1

# Column indexes (1-based) in the stock_in_hand worksheet.
COL = {
    "category": 1,   # "Profit Center"
    "brand": 3,      # "Brand "
    "sku": 4,        # "SKU"
    "description": 5,  # "SKU Description"
    "qty": 7,        # "QTY"
    "consumer_price": 10,  # "Consumer Price"
}

# Profit Center (Excel) -> (slug, display label).
# Order here is the display order used for filters and category groups.
CATEGORY_MAP = [
    ("ACID", "cleaning", "Cleaning"),
    ("BEVERAGES", "beverages", "Beverages"),
    ("CONFECTIONERY", "confectionery", "Confectionery"),
    ("COOKING", "cooking", "Cooking"),
    ("COOKING OIL", "cooking-oil", "Cooking Oil"),
    ("DYPERS", "diapers", "Diapers"),
    ("FEEDER", "baby-care", "Baby Care"),
    ("FOODS", "foods", "Foods"),
    ("GROCERY", "grocery", "Grocery"),
    ("HAIR CARE", "hair-care", "Hair Care"),
    ("HOME CARE", "home-care", "Home Care"),
    ("LAUNDRY", "laundry", "Laundry"),
    ("MILK", "milk-dairy", "Milk & Dairy"),
    ("ORAL CARE", "oral-care", "Oral Care"),
    ("OTHERS", "others", "Others"),
    ("PP", "household-paper", "Household & Paper"),
    ("RAZOR", "razors", "Razors"),
    ("SKIN CARE", "skin-care", "Skin Care"),
    ("SKIN CLEANSING", "skin-cleansing", "Soap & Cleansing"),
    ("SNAKS", "snacks", "Snacks"),
    ("SPICES", "spices", "Spices"),
    ("STATIONARY", "stationery", "Stationery"),
]

CATEGORY_BY_NAME = {name: (slug, label) for name, slug, label in CATEGORY_MAP}

# Anything that looks like a price on the description, e.g.
# "RS 160", "Rs.160", "RS:50", "RS100", "RS :50", "Rs 10 Rs 0".
PRICE_TOKEN_RE = re.compile(r"\bRs\.?\s*:?\s*\d+\b", re.IGNORECASE)
# Redundant "consumer price" marker, e.g. "CP 550", trailing "CP".
CP_TOKEN_RE = re.compile(r"\bCP\s*\d*\b", re.IGNORECASE)
# Junk whitespace between tokens.
SPACE_RE = re.compile(r"\s{2,}")


def slugify(value):
    """Turn a product name into a kebab-case image filename (no extension)."""
    text = re.sub(r"[^A-Za-z0-9]+", "-", value)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text.lower() or "product"


def clean_name(raw):
    """Strip price/consumer-price text from an Excel description."""
    name = str(raw or "").strip()
    name = PRICE_TOKEN_RE.sub(" ", name)
    name = CP_TOKEN_RE.sub(" ", name)
    name = SPACE_RE.sub(" ", name).strip(" -")
    return name


def parse_qty(raw):
    """Return stock as a non-negative integer (1 when the cell is empty)."""
    if raw is None or raw == "":
        return 1  # empty stock defaults to 1
    try:
        qty = int(raw)
    except (TypeError, ValueError):
        return 1
    return max(0, qty)  # negative stock means out of stock


def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_full = os.path.join(project_root, input_path)

    workbook = openpyxl.load_workbook(input_full, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {input_full}.")
    ws = workbook[SHEET_NAME]

    products = []
    seen = set()

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        category_raw = row[COL["category"] - 1]
        description = row[COL["description"] - 1]
        if not category_raw or not description:
            continue

        price = row[COL["consumer_price"] - 1]
        try:
            price = int(price)
        except (TypeError, ValueError):
            continue
        if price <= 0:
            # Unpublishable rows (no price / junk data). Keep the Excel file
            # untouched; just don't expose them in the catalogue.
            continue

        sku = row[COL["sku"] - 1]
        sku = str(sku) if sku is not None else ""
        brand = str(row[COL["brand"] - 1] or "").strip()

        category = CATEGORY_BY_NAME.get(str(category_raw).strip())
        if category is None:
            continue
        category_slug = category[0]

        name = clean_name(description)
        if not name:
            name = brand

        stock = parse_qty(row[COL["qty"] - 1])

        # Skip exact duplicate rows (identical SKU + description).
        dup_key = (sku, str(description).strip())
        if dup_key in seen:
            continue
        seen.add(dup_key)

        products.append(
            {
                "id": len(products) + 1,
                "name": name,
                "category": category_slug,
                "brand": brand,
                "sku": sku,
                "price": price,
                "stock": stock,
                "image": f"images/{slugify(name)}.webp",
            }
        )

    # Make image filenames unique (e.g. two products with the same name).
    image_counts = {}
    for product in products:
        image = product["image"]
        count = image_counts.get(image, 0)
        if count:
            base, ext = image.rsplit(".", 1)
            product["image"] = f"{base}-{count}.{ext}"
        image_counts[image] = image_counts.get(image, 0) + 1

    output_path = os.path.join(project_root, "products.json")
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(products, fh, ensure_ascii=False, indent=2)

    print(f"Imported {len(products)} products -> {output_path}")


if __name__ == "__main__":
    main()
